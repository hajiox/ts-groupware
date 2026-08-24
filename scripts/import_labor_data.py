"""
Import labor-office payroll files under C:\\作業用\\労務 into the TSG payroll schema.

Default is dry-run. Use --apply only after reviewing the JSON summary.
Requires Python packages available in the Codex bundled runtime: openpyxl and pypdf.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from pypdf import PdfReader
except Exception as exc:  # pragma: no cover - setup guard
    print(f"Missing parser dependency: {exc}", file=sys.stderr)
    print("Use the Codex bundled Python runtime or install openpyxl and pypdf.", file=sys.stderr)
    raise


DEFAULT_SOURCE_ROOT = Path("C:/作業用/労務")
DEFAULT_ATTENDANCE_MONTH = "2026-05-01"
DEFAULT_PAYROLL_MONTH = "2026-06-01"
DEFAULT_PAY_DATE = "2026-06-10"
ATTENDANCE_ROUNDING_UNIT_MINUTES = 15

WORKPLACE_CODES = {
    "本社": "hq",
    "会津ブランド館": "aizu_brandhall",
    "ブランド館": "aizu_brandhall",
    "道の駅": "michinoeki",
    "食のブランド館": "food_brandhall",
    "会津しこん": "aizu_shikon",
    "しこん": "aizu_shikon",
}

PAYROLL_KIND_LABELS = {
    "monthly": "monthly payroll",
    "bonus": "bonus payroll",
    "adjustment": "payroll adjustment",
}


def compact_name(value: str | None, *, remove_usage_markers: bool = True) -> str:
    text = value or ""
    text = re.sub(r"[\s　]+", "", text)
    if remove_usage_markers:
        text = re.sub(r"[（(](通勤|業務)[）)]", "", text)
    return text.strip()


def parse_date(value: str | None) -> str | None:
    if not value:
        return None
    value = value.strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def month_end(month_start: str) -> str:
    year, month, _ = [int(part) for part in month_start.split("-")]
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    return date.fromordinal(next_month.toordinal() - 1).isoformat()


def add_months(month_start: str, months: int) -> str:
    year, month, _ = [int(part) for part in month_start.split("-")]
    total_month = (year * 12 + (month - 1)) + months
    next_year = total_month // 12
    next_month = total_month % 12 + 1
    return date(next_year, next_month, 1).isoformat()


def payroll_month_from_folder_name(folder_name: str) -> str | None:
    match = re.search(r"(20\d{2})[.\-_年/ ]\s*(\d{1,2})", folder_name)
    if match:
        return date(int(match.group(1)), int(match.group(2)), 1).isoformat()

    reiwa_match = re.search(r"R0?(\d{1,2})[.\-_年/ ]\s*(\d{1,2})", folder_name, re.IGNORECASE)
    if reiwa_match:
        year = 2018 + int(reiwa_match.group(1))
        return date(year, int(reiwa_match.group(2)), 1).isoformat()

    return None


def pay_date_for_month(payroll_month: str) -> str:
    year, month, _ = [int(part) for part in payroll_month.split("-")]
    return date(year, month, 10).isoformat()


def payroll_kind_from_folder_name(folder_name: str) -> str:
    if "賞与" in folder_name:
        return "bonus"
    return "monthly"


def compact_text(value: str | None) -> str:
    return re.sub(r"[\s\u3000]+", "", value or "")


def normalize_employee_code(value: str | int | None) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", "", str(value))
    if not text:
        return None
    if text.isdigit():
        return str(int(text))
    return text


def parse_japanese_pay_date(text: str, payroll_month: str) -> str | None:
    compact = compact_text(text)
    match = re.search(r"支給日[:：](?:(20\d{2})/)?(\d{1,2})/(\d{1,2})", compact)
    if not match:
        return None
    default_year, _, _ = [int(part) for part in payroll_month.split("-")]
    year = int(match.group(1) or default_year)
    month = int(match.group(2))
    day = int(match.group(3))
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def read_env_file(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        value = value.strip().strip('"').strip("'")
        env[key.strip()] = value
    return env


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def guess_document_type(relative_path: str) -> str:
    name = Path(relative_path).name
    if "賞与一覧" in name:
        return "bonus_statement"
    if "従業員一覧" in name:
        return "employee_master"
    if "休憩時間" in name:
        return "break_rules"
    if "時給一覧" in name:
        return "hourly_rates"
    if "通勤費" in name:
        return "commute"
    if "給与計算チェックリスト" in name:
        return "payroll_checklist"
    if "勤怠チェックリスト" in name:
        return "attendance_checklist"
    if "支給控除一覧" in name:
        return "payroll_statement"
    if "事業所負担保険料" in name:
        return "employer_insurance"
    if "賃金台帳" in name:
        return "wage_ledger"
    if "給与集計" in name:
        return "workplace_payroll_summary"
    if "確認資料" in relative_path or name.startswith("R8.6"):
        return "timecard_source_scan"
    if "勤怠一覧" in name or "従業員毎勤怠集計" in name:
        return "attendance_source"
    return "unknown"


def convert_xls_to_xlsx(path: Path) -> Path | None:
    output = Path(tempfile.gettempdir()) / f"tsg_labor_{path.stem}.xlsx"
    input_path = path.resolve()
    output_path = output.resolve()
    ps = f"""
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
  $wb = $excel.Workbooks.Open('{str(input_path).replace("'", "''")}', 0, $true)
  $wb.SaveAs('{str(output_path).replace("'", "''")}', 51)
  $wb.Close($false)
}} finally {{
  $excel.Quit()
}}
"""
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=90,
        )
    except Exception as exc:
        print(f"Could not convert xls via Excel COM: {path} ({exc})", file=sys.stderr)
        return None
    return output if output.exists() else None


def extract_excel_rows(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    actual_path = path
    converted_from_xls = False
    if path.suffix.lower() == ".xls":
        converted = convert_xls_to_xlsx(path)
        if not converted:
            return [], {"status": "failed", "error": "xls conversion failed"}
        actual_path = converted
        converted_from_xls = True

    wb = openpyxl.load_workbook(actual_path, data_only=False)
    rows: list[dict[str, Any]] = []
    formula_count = 0
    nonempty_count = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            cells = []
            values = []
            has_value = False
            for cell in row:
                value = safe_json_value(cell.value)
                if value is not None:
                    has_value = True
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1
                cells.append(
                    {
                        "coordinate": cell.coordinate,
                        "column": cell.column,
                        "value": value,
                        "is_formula": isinstance(value, str) and value.startswith("="),
                    }
                )
                values.append(value)
            if has_value:
                nonempty_count += 1
                rows.append(
                    {
                        "row_kind": "excel_row",
                        "sheet_name": sheet.title,
                        "row_index": row[0].row,
                        "row_data": {"values": values, "cells": cells},
                        "raw_text": " | ".join(str(v) for v in values if v is not None),
                    }
                )
    return rows, {
        "status": "extracted",
        "sheets": [sheet.title for sheet in wb.worksheets],
        "nonempty_rows": nonempty_count,
        "formula_count": formula_count,
        "converted_from_xls": converted_from_xls,
    }


def extract_pdf_rows(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    reader = PdfReader(str(path))
    extracted_pages = 0
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        normalized = re.sub(r"\s+", " ", text).strip()
        if normalized:
            extracted_pages += 1
        rows.append(
            {
                "row_kind": "pdf_page",
                "page_number": index,
                "row_index": index,
                "row_data": {"text_length": len(text), "preview": normalized[:500]},
                "raw_text": text,
            }
        )
    status = "extracted" if extracted_pages else "image_only"
    return rows, {"status": status, "pages": len(reader.pages), "extracted_pages": extracted_pages}


def collect_documents(source_root: Path, attendance_month: str, payroll_month: str, payroll_kind: str) -> list[dict[str, Any]]:
    documents: list[dict[str, Any]] = []
    for path in sorted(source_root.rglob("*")):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix not in {".pdf", ".xlsx", ".xls"}:
            continue
        relative = str(path.relative_to(source_root))
        doc_type = guess_document_type(relative)
        rows: list[dict[str, Any]] = []
        summary: dict[str, Any]
        try:
            if suffix in {".xlsx", ".xls"}:
                rows, summary = extract_excel_rows(path)
            else:
                rows, summary = extract_pdf_rows(path)
        except Exception as exc:
            rows, summary = [], {"status": "failed", "error": str(exc)}

        documents.append(
            {
                "relative_path": relative,
                "file_name": path.name,
                "file_extension": suffix.lstrip("."),
                "file_size": path.stat().st_size,
                "sha256": sha256_file(path),
                "document_type": doc_type,
                "payroll_kind": payroll_kind,
                "target_attendance_month": attendance_month,
                "target_payroll_month": payroll_month,
                "extraction_status": summary.get("status", "partial"),
                "extraction_notes": summary.get("error"),
                "extracted_summary": summary,
                "rows": rows,
            }
        )
    return documents


def extract_employee_master(documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    employees: list[dict[str, Any]] = []
    pattern = re.compile(
        r"^(?P<code>\d+)\s+(?P<name>.+?)\s+(?P<kana>[ｦ-ﾟA-Za-z\s]+)\s+"
        r"(?P<birth>\d{4}/\d{1,2}/\d{1,2})\s+(?P<hire>\d{4}/\d{1,2}/\d{1,2})$"
    )
    for document in documents:
        if document["document_type"] != "employee_master":
            continue
        for row in document["rows"]:
            text = row.get("raw_text") or ""
            for line in text.splitlines():
                match = pattern.match(line.strip())
                if not match:
                    continue
                raw_name = match.group("name").strip()
                gender = "unknown"
                if "（男性）" in raw_name or "(男性)" in raw_name:
                    gender = "male"
                elif "（女性）" in raw_name or "(女性)" in raw_name:
                    gender = "female"
                employees.append(
                    {
                        "employee_code": normalize_employee_code(match.group("code")),
                        "display_name": raw_name,
                        "real_name": raw_name,
                        "kana": match.group("kana").strip(),
                        "birth_date": parse_date(match.group("birth")),
                        "hire_date": parse_date(match.group("hire")),
                        "gender": gender,
                        "payroll_status": "active",
                        "raw_payload": {"source_path": document["relative_path"], "raw_line": line},
                    }
                )
    return employees


def find_document(documents: list[dict[str, Any]], file_name: str) -> dict[str, Any] | None:
    for document in documents:
        if document["file_name"] == file_name:
            return document
    return None


def workbook_from_source(source_root: Path, file_name: str) -> openpyxl.Workbook | None:
    path = source_root / file_name
    if not path.exists():
        return None
    return openpyxl.load_workbook(path, data_only=False)


def extract_commute_data(source_root: Path, documents: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    wb = workbook_from_source(source_root, "通勤費R8.xlsx")
    if not wb:
        return {"routes": [], "monthly_results": [], "extra_employees": []}
    doc = find_document(documents, "通勤費R8.xlsx")
    routes: list[dict[str, Any]] = []
    monthly: list[dict[str, Any]] = []
    extra_employees: list[dict[str, Any]] = []

    if "通勤距離" in wb.sheetnames:
        ws = wb["通勤距離"]
        for row_index in range(4, ws.max_row + 1):
            name = ws.cell(row_index, 1).value
            if not name:
                continue
            raw_name = str(name).strip()
            tax_free_limit = safe_json_value(ws.cell(row_index, 5).value)
            extra_employees.append({"display_name": raw_name, "real_name": raw_name, "raw_payload": {"source": "commute"}})
            for col, workplace_code in [(2, "hq"), (3, "michinoeki"), (4, "aizu_brandhall")]:
                distance = ws.cell(row_index, col).value
                if distance in (None, ""):
                    continue
                route_type = "business" if re.search(r"[（(]業務[）)]", raw_name) else "commute"
                round_trip_multiplier = 1 if "藤田" in raw_name else 2
                routes.append(
                    {
                        "name": raw_name,
                        "workplace_code": workplace_code,
                        "route_type": route_type,
                        "one_way_distance_km": float(distance),
                        "round_trip_multiplier": round_trip_multiplier,
                        "yen_per_km": 16,
                        "tax_free_limit": float(tax_free_limit) if tax_free_limit not in (None, "") else None,
                        "monthly_cap": 10000,
                        "effective_from": "2026-05-01",
                        "source_relative_path": doc["relative_path"] if doc else None,
                        "note": "通勤費R8.xlsx から取込",
                        "raw_payload": {
                            "sheet": "通勤距離",
                            "row": row_index,
                            "raw_name": raw_name,
                            "tax_free_limit": tax_free_limit,
                        },
                    }
                )

    if "R8.5" in wb.sheetnames:
        ws = wb["R8.5"]
        for row_index in range(6, ws.max_row + 1):
            raw_name = ws.cell(row_index, 1).value
            if not raw_name:
                continue
            raw_name = str(raw_name).strip()
            extra_employees.append({"display_name": raw_name, "real_name": raw_name, "raw_payload": {"source": "commute_monthly"}})
            for col, workplace_code in [(2, "hq"), (3, "michinoeki"), (4, "aizu_brandhall")]:
                days = ws.cell(row_index, col).value
                if days in (None, ""):
                    continue
                monthly.append(
                    {
                        "name": raw_name,
                        "workplace_code": workplace_code,
                        "work_days": float(days),
                        "total_amount": 0,
                        "tax_free_amount": 0,
                        "taxable_amount": 0,
                        "source_relative_path": doc["relative_path"] if doc else None,
                        "calculation_snapshot": {
                            "sheet": "R8.5",
                            "row": row_index,
                            "days_cell": ws.cell(row_index, col).coordinate,
                            "formula_cells": {
                                "regular": safe_json_value(ws.cell(row_index, 7).value),
                                "michinoeki": safe_json_value(ws.cell(row_index, 8).value),
                                "brandhall": safe_json_value(ws.cell(row_index, 9).value),
                                "total": safe_json_value(ws.cell(row_index, 10).value),
                                "tax_free": safe_json_value(ws.cell(row_index, 11).value),
                                "taxable": safe_json_value(ws.cell(row_index, 12).value),
                            },
                        },
                    }
                )

    return {"routes": routes, "monthly_results": monthly, "extra_employees": extra_employees}


def extract_payroll_checklists(source_root: Path, documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    wb = workbook_from_source(source_root, "給与計算チェックリスト.xlsx")
    if not wb:
        return []
    doc = find_document(documents, "給与計算チェックリスト.xlsx")
    items: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_index in range(5, min(ws.max_row, 20) + 1):
            workplace = ws.cell(row_index, 1).value
            if not workplace or "／" in str(workplace):
                continue
            workplace_name = str(workplace).strip()
            if workplace_name not in WORKPLACE_CODES:
                continue
            target_count = ws.cell(row_index, 6).value
            calculated_count = ws.cell(row_index, 7).value
            items.append(
                {
                    "workplace_name": workplace_name,
                    "checklist_type": "payroll_calculation",
                    "input_by": safe_json_value(ws.cell(row_index, 3).value),
                    "confirmed_by": safe_json_value(ws.cell(row_index, 5).value),
                    "target_employee_count": int(target_count) if isinstance(target_count, (int, float)) else None,
                    "calculated_employee_count": int(calculated_count) if isinstance(calculated_count, (int, float)) else None,
                    "source_relative_path": doc["relative_path"] if doc else None,
                    "raw_payload": {"sheet": sheet_name, "row": row_index},
                }
            )
    return items


def aggregate_commute_monthly_results(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (compact_name(row.get("name"), remove_usage_markers=True), row["workplace_code"])
        current = grouped.get(key)
        if not current:
            current = {
                **row,
                "name": key[0],
                "work_days": 0,
                "calculation_snapshot": {"rows": []},
            }
            grouped[key] = current
        current["work_days"] += float(row.get("work_days") or 0)
        current["calculation_snapshot"]["rows"].append(row.get("calculation_snapshot", {}))
    return list(grouped.values())


def extract_cost_allocations(documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    allocations: list[dict[str, Any]] = []
    for document in documents:
        if document["document_type"] != "workplace_payroll_summary":
            continue
        for row in document["rows"]:
            if row.get("sheet_name") not in {"各事業所", "6月支給合計額"}:
                continue
            values = (row.get("row_data") or {}).get("values") or []
            text_values = [str(v).strip() if v is not None else "" for v in values]
            first = text_values[0] if text_values else ""
            if not first or ("計" not in first and "総支給額" not in first):
                continue
            workplace_name = first.replace("計", "").strip()
            if first == "総支給額":
                workplace_name = None
                allocation_key = "総支給額"
            elif workplace_name in WORKPLACE_CODES:
                allocation_key = first
            else:
                continue
            amount = None
            for value in text_values[1:]:
                if re.match(r"^-?[\d,]+(\.\d+)?$", value):
                    amount = float(value.replace(",", ""))
                    break
            if amount is None:
                continue
            allocations.append(
                {
                    "workplace_name": workplace_name,
                    "allocation_key": allocation_key,
                    "amount": amount,
                    "source_relative_path": document["relative_path"],
                    "raw_payload": {
                        "sheet": row.get("sheet_name"),
                        "row": row.get("row_index"),
                        "values": values,
                    },
                }
            )
    return allocations


AMOUNT_TOKEN_RE = re.compile(r"-?\s*(?:\d\s*){1,3}(?:,\s*(?:\d\s*){3})+")

PAYROLL_RESULT_FIELDS = {
    "taxable_payment_total": {"labels": ["課税支給合計"], "exclude": ["非課税支給合計"]},
    "non_taxable_payment_total": {"labels": ["非課税支給合計"], "exclude": []},
    "payment_total": {"labels": ["支給合計"], "exclude": ["課税支給合計", "非課税支給合計"]},
    "social_insurance_total": {"labels": ["社保控除合計"], "exclude": []},
    "deduction_total": {"labels": ["控除合計"], "exclude": ["社保控除合計", "その他控除合計"]},
    "taxable_income": {"labels": ["課税対象額"], "exclude": ["前月課税対象額"]},
    "net_payment": {"labels": ["差引支給額"], "exclude": []},
    "cash_payment": {"labels": ["現金支給額"], "exclude": []},
    "transfer_payment": {"labels": ["振込支給額", "振込c支給額"], "exclude": []},
}


def parse_amount_token(token: str) -> int | None:
    digits = re.sub(r"[^\d-]", "", token)
    if not digits or digits == "-":
        return None
    return int(digits)


def extract_amount_tokens(value: str) -> list[int]:
    amounts: list[int] = []
    for match in AMOUNT_TOKEN_RE.finditer(value):
        amount = parse_amount_token(match.group(0))
        if amount is not None:
            amounts.append(amount)
    return amounts


def line_amounts_for_label(line: str, label: str, employee_count: int) -> list[int] | None:
    compact_line = compact_text(line)
    compact_label = compact_text(label)
    if compact_label not in compact_line:
        return None
    before, after = compact_line.split(compact_label, 1)
    before_amounts = extract_amount_tokens(before)
    after_amounts = extract_amount_tokens(after)
    if employee_count <= 1:
        amount = before_amounts[-1] if before_amounts else (after_amounts[0] if after_amounts else 0)
        return [amount]

    amounts = [before_amounts[-1] if before_amounts else 0]
    amounts.extend(after_amounts[: employee_count - 1])
    if len(amounts) < employee_count:
        amounts.extend([0] * (employee_count - len(amounts)))
    return amounts[:employee_count]


def extract_field_amounts(lines: list[str], employee_count: int, field: str) -> list[int]:
    config = PAYROLL_RESULT_FIELDS[field]
    for line in lines:
        compact_line = compact_text(line)
        if any(compact_text(excluded) in compact_line for excluded in config["exclude"]):
            continue
        for label in config["labels"]:
            amounts = line_amounts_for_label(line, label, employee_count)
            if amounts is not None:
                return amounts
    return [0] * employee_count


def employee_candidates(employees: list[dict[str, Any]]) -> list[dict[str, str | None]]:
    candidates: list[dict[str, str | None]] = []
    seen: set[tuple[str, str | None]] = set()
    for employee in employees:
        name = employee.get("real_name") or employee.get("display_name")
        key = compact_text(compact_name(name))
        if not key:
            continue
        marker = (key, employee.get("employee_code"))
        if marker in seen:
            continue
        seen.add(marker)
        candidates.append(
            {
                "name": name,
                "employee_code": employee.get("employee_code"),
                "key": key,
            }
        )
    candidates.sort(key=lambda item: len(str(item["key"])), reverse=True)
    return candidates


def known_page_employees(text: str, employees: list[dict[str, Any]]) -> list[dict[str, str | None]]:
    header = "\n".join((text or "").splitlines()[:8])
    compact_header = compact_text(header)
    matches: list[tuple[int, dict[str, str | None]]] = []
    occupied: list[tuple[int, int]] = []
    for candidate in employee_candidates(employees):
        key = str(candidate["key"])
        position = compact_header.find(key)
        if position < 0:
            continue
        end = position + len(key)
        if any(not (end <= start or position >= stop) for start, stop in occupied):
            continue
        occupied.append((position, end))
        matches.append((position, candidate))
    matches.sort(key=lambda item: item[0])
    return [match[1] for match in matches[:5]]


def bonus_header_employees(text: str) -> list[dict[str, str | None]]:
    header = "\n".join((text or "").splitlines()[:20])
    compact_header = compact_text(header)
    compact_header = compact_header.split("賞与支給日", 1)[0]
    compact_header = compact_header.split("月賞与分", 1)[0]
    employees: list[dict[str, str | None]] = []
    for match in re.finditer(r"(?P<code>\d{6})(?P<name>.*?)(?=\d{6}|$)", compact_header):
        code = normalize_employee_code(match.group("code"))
        name = re.sub(r"\d+名.*$", "", match.group("name"))
        name = re.sub(r"\d+$", "", name)
        if not name or "株式会社" in name or len(name) > 20:
            continue
        employees.append({"name": name, "employee_code": code, "key": compact_text(name)})
    return employees[:5]


def extract_pay_date_from_documents(documents: list[dict[str, Any]], payroll_month: str) -> str | None:
    for document in documents:
        if document["document_type"] not in {"payroll_statement", "bonus_statement"}:
            continue
        for row in document["rows"]:
            pay_date = parse_japanese_pay_date(row.get("raw_text") or "", payroll_month)
            if pay_date:
                return pay_date
    return None


def extract_payroll_results(documents: list[dict[str, Any]], employees: list[dict[str, Any]], payroll_kind: str) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    seen: set[tuple[str | None, str, str, int]] = set()
    for document in documents:
        if document["document_type"] not in {"payroll_statement", "bonus_statement"}:
            continue
        for row in document["rows"]:
            text = row.get("raw_text") or ""
            lines = text.splitlines()
            page_number = int(row.get("page_number") or row.get("row_index") or 0)
            page_employees = bonus_header_employees(text) if document["document_type"] == "bonus_statement" else []
            if not page_employees:
                page_employees = known_page_employees(text, employees)
            if not page_employees:
                continue

            employee_count = len(page_employees)
            field_amounts = {
                field: extract_field_amounts(lines, employee_count, field)
                for field in PAYROLL_RESULT_FIELDS
            }
            if not any(field_amounts[field][index] for field in ("payment_total", "net_payment") for index in range(employee_count)):
                continue

            for index, employee in enumerate(page_employees):
                name = str(employee.get("name") or "")
                employee_code = employee.get("employee_code")
                key = (employee_code, compact_text(name), document["sha256"], page_number)
                if key in seen:
                    continue
                seen.add(key)
                values = {field: field_amounts[field][index] if index < len(field_amounts[field]) else 0 for field in PAYROLL_RESULT_FIELDS}
                if not values["payment_total"] and not values["net_payment"]:
                    continue
                results.append(
                    {
                        "name": name,
                        "employee_code": employee_code,
                        "payroll_kind": payroll_kind,
                        "source_relative_path": document["relative_path"],
                        "source_sha256": document["sha256"],
                        "page_number": page_number,
                        "taxable_payment_total": values["taxable_payment_total"],
                        "non_taxable_payment_total": values["non_taxable_payment_total"],
                        "payment_total": values["payment_total"],
                        "social_insurance_total": values["social_insurance_total"],
                        "deduction_total": values["deduction_total"],
                        "taxable_income": values["taxable_income"],
                        "net_payment": values["net_payment"],
                        "cash_payment": values["cash_payment"],
                        "transfer_payment": values["transfer_payment"],
                        "raw_payload": {
                            "source_relative_path": document["relative_path"],
                            "page_number": page_number,
                            "employee_index": index,
                            "employee_count_on_page": employee_count,
                            "document_type": document["document_type"],
                            "extraction": "pdf_label_row_amounts",
                        },
                    }
                )
    return results


class SupabaseRest:
    def __init__(self, url: str, service_key: str) -> None:
        self.base = url.rstrip("/")
        self.key = service_key

    def request(self, method: str, path: str, body: Any | None = None, prefer: str | None = None) -> Any:
        data = None
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
        }
        if prefer:
            headers["Prefer"] = prefer
        if body is not None:
            data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(f"{self.base}/rest/v1/{path}", data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                text = response.read().decode("utf-8")
                return json.loads(text) if text else None
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Supabase {method} {path} failed: {exc.code} {detail}") from exc

    def select(self, table: str, query: str) -> list[dict[str, Any]]:
        return self.request("GET", f"{table}?{query}") or []

    def insert(self, table: str, rows: list[dict[str, Any]] | dict[str, Any]) -> list[dict[str, Any]]:
        return self.request("POST", table, rows, prefer="return=representation") or []

    def upsert(self, table: str, rows: list[dict[str, Any]], on_conflict: str) -> list[dict[str, Any]]:
        if not rows:
            return []
        path = f"{table}?on_conflict={urllib.parse.quote(on_conflict)}"
        return self.request("POST", path, rows, prefer="resolution=merge-duplicates,return=representation") or []

    def patch(self, table: str, query: str, values: dict[str, Any]) -> list[dict[str, Any]]:
        return self.request("PATCH", f"{table}?{query}", values, prefer="return=representation") or []

    def delete(self, table: str, query: str) -> None:
        self.request("DELETE", f"{table}?{query}")


def chunked(items: list[dict[str, Any]], size: int = 250) -> list[list[dict[str, Any]]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def normalize_for_user_match(value: str | None) -> str:
    text = compact_name(value, remove_usage_markers=True)
    return text


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", str(value))
    return "'" + text.replace("'", "''") + "'"


def clean_json_value(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    if isinstance(value, list):
        return [clean_json_value(item) for item in value]
    if isinstance(value, dict):
        return {key: clean_json_value(item) for key, item in value.items()}
    return value


def sql_json(value: Any) -> str:
    return sql_literal(json.dumps(clean_json_value(value or {}), ensure_ascii=False)) + "::jsonb"


def sql_date(value: str | None) -> str:
    return sql_literal(value) if value else "NULL"


def sql_employee_match(name: str) -> str:
    key = normalize_for_user_match(name)
    return (
        "(select id from gw_payroll_employees "
        "where regexp_replace(coalesce(real_name, display_name), '[\\s　]+', '', 'g') = "
        f"{sql_literal(key)} order by employee_code nulls last, created_at limit 1)"
    )


def sql_employee_match_result(result: dict[str, Any]) -> str:
    by_name = sql_employee_match(result["name"])
    if result.get("employee_code"):
        return (
            "coalesce("
            f"(select id from gw_payroll_employees where employee_code = {sql_literal(result['employee_code'])} limit 1), "
            f"{by_name})"
        )
    return by_name


def payroll_period_where(preview: dict[str, Any], table_alias: str = "period") -> str:
    return (
        f"{table_alias}.payroll_month = {sql_date(preview['payroll_month'])} "
        f"AND {table_alias}.payroll_kind = {sql_literal(preview.get('payroll_kind', 'monthly'))}"
    )


def generate_data_sql(preview: dict[str, Any]) -> str:
    batch_id = str(uuid.uuid4())
    lines: list[str] = [
        "-- Generated by scripts/import_labor_data.py",
        "BEGIN;",
        "SET LOCAL statement_timeout = '120s';",
        "",
        "INSERT INTO gw_labor_import_batches (id, source_root, payroll_kind, target_attendance_month, target_payroll_month, period_start, period_end, pay_date, summary)",
        (
            "VALUES ("
            f"{sql_literal(batch_id)}::uuid, {sql_literal(preview['source_root'])}, "
            f"{sql_literal(preview.get('payroll_kind', 'monthly'))}, "
            f"{sql_date(preview['attendance_month'])}, {sql_date(preview['payroll_month'])}, "
            f"{sql_date(preview['period_start'])}, {sql_date(preview['period_end'])}, {sql_date(preview['pay_date'])}, "
            f"{sql_json(preview['summary'])})"
        ),
        "ON CONFLICT (id) DO NOTHING;",
        "",
        "INSERT INTO gw_attendance_periods (attendance_month, period_start, period_end, cutoff_day, rounding_unit_minutes, break_rule_set)",
        (
            "VALUES ("
            f"{sql_date(preview['attendance_month'])}, {sql_date(preview['period_start'])}, {sql_date(preview['period_end'])}, "
            f"{int(preview['period_end'].split('-')[-1])}, {ATTENDANCE_ROUNDING_UNIT_MINUTES}, 'default')"
        ),
        f"ON CONFLICT (attendance_month) DO UPDATE SET period_start = excluded.period_start, period_end = excluded.period_end, rounding_unit_minutes = {ATTENDANCE_ROUNDING_UNIT_MINUTES}, break_rule_set = 'default', updated_at = now();",
        "",
        "INSERT INTO gw_payroll_periods (payroll_month, payroll_kind, attendance_period_id, attendance_month, period_start, period_end, pay_date)",
        (
            "SELECT "
            f"{sql_date(preview['payroll_month'])}, {sql_literal(preview.get('payroll_kind', 'monthly'))}, id, {sql_date(preview['attendance_month'])}, "
            f"{sql_date(preview['period_start'])}, {sql_date(preview['period_end'])}, {sql_date(preview['pay_date'])} "
            "FROM gw_attendance_periods "
            f"WHERE attendance_month = {sql_date(preview['attendance_month'])}"
        ),
        "ON CONFLICT (payroll_month, payroll_kind) DO UPDATE SET attendance_period_id = excluded.attendance_period_id, attendance_month = excluded.attendance_month, period_start = excluded.period_start, period_end = excluded.period_end, pay_date = excluded.pay_date, updated_at = now();",
        "",
    ]

    for document in preview["documents"]:
        doc_values = {
            "import_batch_id": batch_id,
            "relative_path": document["relative_path"],
            "file_name": document["file_name"],
            "file_extension": document["file_extension"],
            "file_size": document["file_size"],
            "sha256": document["sha256"],
            "document_type": document["document_type"],
            "target_attendance_month": document["target_attendance_month"],
            "target_payroll_month": document["target_payroll_month"],
            "extraction_status": document["extraction_status"],
            "extraction_notes": document.get("extraction_notes"),
            "extracted_summary": document["extracted_summary"],
        }
        lines.extend(
            [
                f"-- Source document: {document['relative_path']}",
                "INSERT INTO gw_labor_source_documents (import_batch_id, relative_path, file_name, file_extension, file_size, sha256, document_type, target_attendance_month, target_payroll_month, extraction_status, extraction_notes, extracted_summary)",
                (
                    "VALUES ("
                    f"{sql_literal(doc_values['import_batch_id'])}::uuid, {sql_literal(doc_values['relative_path'])}, "
                    f"{sql_literal(doc_values['file_name'])}, {sql_literal(doc_values['file_extension'])}, {doc_values['file_size']}, "
                    f"{sql_literal(doc_values['sha256'])}, {sql_literal(doc_values['document_type'])}, "
                    f"{sql_date(doc_values['target_attendance_month'])}, {sql_date(doc_values['target_payroll_month'])}, "
                    f"{sql_literal(doc_values['extraction_status'])}, {sql_literal(doc_values['extraction_notes'])}, {sql_json(doc_values['extracted_summary'])})"
                ),
                "ON CONFLICT (sha256) DO UPDATE SET import_batch_id = excluded.import_batch_id, relative_path = excluded.relative_path, file_name = excluded.file_name, file_extension = excluded.file_extension, file_size = excluded.file_size, document_type = excluded.document_type, target_attendance_month = excluded.target_attendance_month, target_payroll_month = excluded.target_payroll_month, extraction_status = excluded.extraction_status, extraction_notes = excluded.extraction_notes, extracted_summary = excluded.extracted_summary, updated_at = now();",
                f"DELETE FROM gw_labor_source_rows WHERE source_document_id = (select id from gw_labor_source_documents where sha256 = {sql_literal(document['sha256'])});",
            ]
        )
        for chunk in chunked(document["rows"], 40):
            values = []
            for row in chunk:
                values.append(
                    "("
                    f"(select id from gw_labor_source_documents where sha256 = {sql_literal(document['sha256'])}), "
                    f"{sql_literal(row.get('row_kind', 'row'))}, "
                    f"{sql_literal(row.get('sheet_name'))}, "
                    f"{sql_literal(row.get('page_number'))}, "
                    f"{sql_literal(row.get('row_index', 0))}, "
                    f"{sql_json(row.get('row_data'))}, "
                    f"{sql_literal(row.get('raw_text'))}"
                    ")"
                )
            lines.append("INSERT INTO gw_labor_source_rows (source_document_id, row_kind, sheet_name, page_number, row_index, row_data, raw_text)")
            lines.append("VALUES\n" + ",\n".join(values) + ";")
        lines.append("")

    for employee in preview["employees"]:
        name_key = normalize_for_user_match(employee.get("real_name") or employee.get("display_name"))
        raw_payload = employee.get("raw_payload", {})
        source_is_payroll_result = raw_payload.get("source") == "payroll_result"
        display_name_expr = (
            f"CASE WHEN display_name IS NULL OR display_name = '' THEN {sql_literal(employee.get('display_name'))} ELSE display_name END"
            if source_is_payroll_result
            else sql_literal(employee.get("display_name"))
        )
        real_name_expr = (
            f"CASE WHEN real_name IS NULL OR real_name = '' THEN {sql_literal(employee.get('real_name') or employee.get('display_name'))} ELSE real_name END"
            if source_is_payroll_result
            else sql_literal(employee.get("real_name") or employee.get("display_name"))
        )
        lines.extend(
            [
                "DO $$",
                "DECLARE",
                "  v_user_id uuid;",
                "  v_employee_id uuid;",
                "  v_workplace_id uuid;",
                "BEGIN",
                "  SELECT id INTO v_user_id FROM gw_users",
                f"  WHERE regexp_replace(coalesce(real_name, display_name), '[\\s　]+', '', 'g') = {sql_literal(name_key)}",
                "  ORDER BY CASE WHEN status = 'approved' THEN 0 ELSE 1 END, created_at LIMIT 1;",
                "  IF v_user_id IS NOT NULL THEN",
                "    SELECT id INTO v_workplace_id FROM gw_workplaces",
                "    WHERE code = CASE",
                "      WHEN (SELECT department FROM gw_users WHERE id = v_user_id) = 'フロア' THEN 'aizu_brandhall'",
                "      WHEN (SELECT department FROM gw_users WHERE id = v_user_id) = '道の駅' THEN 'michinoeki'",
                "      ELSE 'hq'",
                "    END;",
                "  END IF;",
            ]
        )
        if employee.get("employee_code"):
            lines.append(f"  SELECT id INTO v_employee_id FROM gw_payroll_employees WHERE employee_code = {sql_literal(employee['employee_code'])} LIMIT 1;")
        lines.extend(
            [
                "  IF v_employee_id IS NULL AND v_user_id IS NOT NULL THEN",
                "    SELECT id INTO v_employee_id FROM gw_payroll_employees WHERE user_id = v_user_id LIMIT 1;",
                "  END IF;",
                "  IF v_employee_id IS NULL THEN",
                "    SELECT id INTO v_employee_id FROM gw_payroll_employees",
                f"    WHERE regexp_replace(coalesce(real_name, display_name), '[\\s　]+', '', 'g') = {sql_literal(name_key)}",
                "    ORDER BY employee_code nulls last, created_at LIMIT 1;",
                "  END IF;",
                "  IF v_employee_id IS NULL THEN",
                "    INSERT INTO gw_payroll_employees (user_id, employee_code, display_name, real_name, kana, birth_date, hire_date, gender, department, default_workplace_id, payroll_status, raw_payload)",
                "    VALUES (",
                "      v_user_id,",
                f"      {sql_literal(employee.get('employee_code'))},",
                f"      {sql_literal(employee.get('display_name'))},",
                f"      {sql_literal(employee.get('real_name') or employee.get('display_name'))},",
                f"      {sql_literal(employee.get('kana'))},",
                f"      {sql_date(employee.get('birth_date'))},",
                f"      {sql_date(employee.get('hire_date'))},",
                f"      {sql_literal(employee.get('gender') or 'unknown')},",
                "      (SELECT department FROM gw_users WHERE id = v_user_id),",
                "      v_workplace_id,",
                f"      {sql_literal(employee.get('payroll_status', 'active'))},",
                f"      {sql_json(raw_payload)}",
                "    );",
                "  ELSE",
                "    UPDATE gw_payroll_employees SET",
                f"      employee_code = COALESCE({sql_literal(employee.get('employee_code'))}, employee_code),",
                f"      display_name = {display_name_expr},",
                f"      real_name = {real_name_expr},",
                f"      kana = COALESCE({sql_literal(employee.get('kana'))}, kana),",
                f"      birth_date = COALESCE({sql_date(employee.get('birth_date'))}, birth_date),",
                f"      hire_date = COALESCE({sql_date(employee.get('hire_date'))}, hire_date),",
                f"      gender = COALESCE({sql_literal(employee.get('gender') or None)}, gender),",
                "      user_id = COALESCE(v_user_id, user_id),",
                "      department = COALESCE((SELECT department FROM gw_users WHERE id = v_user_id), department),",
                "      default_workplace_id = COALESCE(v_workplace_id, default_workplace_id),",
                "      raw_payload = raw_payload || " + sql_json(raw_payload) + ",",
                "      updated_at = now()",
                "    WHERE id = v_employee_id;",
                "  END IF;",
                "END $$;",
                "",
            ]
        )

    if preview.get("payroll_results"):
        result_summary = {
            "payroll_kind": preview.get("payroll_kind", "monthly"),
            "result_count": len(preview["payroll_results"]),
            "payment_total": sum(result.get("payment_total") or 0 for result in preview["payroll_results"]),
            "net_payment": sum(result.get("net_payment") or 0 for result in preview["payroll_results"]),
        }
        lines.extend(
            [
                "INSERT INTO gw_payroll_runs (payroll_period_id, source_import_batch_id, run_number, status, calculation_mode, summary)",
                (
                    "SELECT period.id, "
                    f"{sql_literal(batch_id)}::uuid, 1, 'calculated', 'imported', {sql_json(result_summary)} "
                    "FROM gw_payroll_periods period "
                    f"WHERE {payroll_period_where(preview)}"
                ),
                "ON CONFLICT (payroll_period_id, run_number) DO UPDATE SET source_import_batch_id = excluded.source_import_batch_id, status = excluded.status, calculation_mode = excluded.calculation_mode, summary = excluded.summary, updated_at = now();",
                "",
                "DELETE FROM gw_payroll_employee_results WHERE payroll_period_id = (select id from gw_payroll_periods period where " + payroll_period_where(preview) + ");",
            ]
        )
        for result in preview["payroll_results"]:
            employee_match = sql_employee_match_result(result)
            lines.extend(
                [
                    "INSERT INTO gw_payroll_employee_results (payroll_run_id, payroll_period_id, employee_id, taxable_payment_total, non_taxable_payment_total, payment_total, social_insurance_total, deduction_total, taxable_income, net_payment, cash_payment, transfer_payment, source_document_id, raw_payload)",
                    (
                        "SELECT run.id, period.id, matched.employee_id, "
                        f"{result.get('taxable_payment_total') or 0}, {result.get('non_taxable_payment_total') or 0}, {result.get('payment_total') or 0}, "
                        f"{result.get('social_insurance_total') or 0}, {result.get('deduction_total') or 0}, {result.get('taxable_income') or 0}, "
                        f"{result.get('net_payment') or 0}, {result.get('cash_payment') or 0}, {result.get('transfer_payment') or 0}, "
                        f"doc.id, {sql_json(result.get('raw_payload'))} "
                        "FROM gw_payroll_periods period "
                        "JOIN gw_payroll_runs run ON run.payroll_period_id = period.id AND run.run_number = 1 "
                        f"CROSS JOIN LATERAL (SELECT {employee_match} AS employee_id) matched "
                        f"LEFT JOIN gw_labor_source_documents doc ON doc.sha256 = {sql_literal(result.get('source_sha256'))} "
                        f"WHERE {payroll_period_where(preview)} AND matched.employee_id IS NOT NULL"
                    ),
                    "ON CONFLICT (payroll_run_id, employee_id) DO UPDATE SET taxable_payment_total = excluded.taxable_payment_total, non_taxable_payment_total = excluded.non_taxable_payment_total, payment_total = excluded.payment_total, social_insurance_total = excluded.social_insurance_total, deduction_total = excluded.deduction_total, taxable_income = excluded.taxable_income, net_payment = excluded.net_payment, cash_payment = excluded.cash_payment, transfer_payment = excluded.transfer_payment, source_document_id = excluded.source_document_id, raw_payload = excluded.raw_payload, updated_at = now();",
                ]
            )
        lines.append("")

    for route in preview["commute_routes"]:
        employee_match = sql_employee_match(route["name"])
        lines.extend(
            [
                "INSERT INTO gw_commute_routes (employee_id, workplace_id, route_type, one_way_distance_km, round_trip_multiplier, yen_per_km, tax_free_limit, monthly_cap, effective_from, source_document_id, note, raw_payload)",
                (
                    "SELECT "
                    f"{employee_match}, workplace.id, {sql_literal(route['route_type'])}, {route['one_way_distance_km']}, "
                    f"{route['round_trip_multiplier']}, {route['yen_per_km']}, {sql_literal(route.get('tax_free_limit'))}, "
                    f"{route['monthly_cap']}, {sql_date(route.get('effective_from'))}, "
                    f"(select id from gw_labor_source_documents where relative_path = {sql_literal(route.get('source_relative_path'))} order by created_at desc limit 1), "
                    f"{sql_literal(route.get('note'))}, {sql_json(route.get('raw_payload'))} "
                    "FROM gw_workplaces workplace "
                    f"WHERE workplace.code = {sql_literal(route['workplace_code'])} AND {employee_match} IS NOT NULL"
                ),
                "ON CONFLICT (employee_id, workplace_id, route_type, effective_from) DO UPDATE SET one_way_distance_km = excluded.one_way_distance_km, round_trip_multiplier = excluded.round_trip_multiplier, yen_per_km = excluded.yen_per_km, tax_free_limit = excluded.tax_free_limit, monthly_cap = excluded.monthly_cap, source_document_id = excluded.source_document_id, note = excluded.note, raw_payload = excluded.raw_payload;",
            ]
        )

    lines.append(
        "DELETE FROM gw_commute_monthly_results WHERE payroll_period_id = (select id from gw_payroll_periods period where "
        + payroll_period_where(preview)
        + ");"
    )
    for result in preview["commute_monthly_results"]:
        employee_match = sql_employee_match(result["name"])
        source_path = result.get("source_relative_path")
        lines.extend(
            [
                "INSERT INTO gw_commute_monthly_results (payroll_period_id, employee_id, workplace_id, work_days, total_amount, tax_free_amount, taxable_amount, calculation_snapshot, source_document_id)",
                (
                    "SELECT period.id, "
                    f"{employee_match}, workplace.id, {result['work_days']}, {result['total_amount']}, {result['tax_free_amount']}, {result['taxable_amount']}, "
                    f"{sql_json(result.get('calculation_snapshot'))}, "
                    f"(select id from gw_labor_source_documents where relative_path = {sql_literal(source_path)} order by created_at desc limit 1) "
                    "FROM gw_payroll_periods period, gw_workplaces workplace "
                    f"WHERE {payroll_period_where(preview)} "
                    f"AND workplace.code = {sql_literal(result['workplace_code'])} AND {employee_match} IS NOT NULL;"
                ),
            ]
        )

    lines.append(
        "DELETE FROM gw_payroll_checklists WHERE payroll_period_id = (select id from gw_payroll_periods period where "
        + payroll_period_where(preview)
        + ");"
    )
    for checklist in preview["payroll_checklists"]:
        lines.extend(
            [
                "INSERT INTO gw_payroll_checklists (payroll_period_id, source_document_id, workplace_id, checklist_type, input_by, confirmed_by, target_employee_count, calculated_employee_count, raw_payload)",
                (
                    "SELECT period.id, "
                    f"(select id from gw_labor_source_documents where relative_path = {sql_literal(checklist.get('source_relative_path'))} order by created_at desc limit 1), "
                    "workplace.id, "
                    f"{sql_literal(checklist.get('checklist_type'))}, {sql_literal(checklist.get('input_by'))}, {sql_literal(checklist.get('confirmed_by'))}, "
                    f"{sql_literal(checklist.get('target_employee_count'))}, {sql_literal(checklist.get('calculated_employee_count'))}, {sql_json(checklist.get('raw_payload'))} "
                    "FROM gw_payroll_periods period "
                    "LEFT JOIN gw_workplaces workplace ON workplace.name = " + sql_literal(checklist.get("workplace_name")) + " "
                    f"WHERE {payroll_period_where(preview)};"
                ),
            ]
        )

    lines.append(
        "DELETE FROM gw_payroll_cost_allocations WHERE payroll_period_id = (select id from gw_payroll_periods period where "
        + payroll_period_where(preview)
        + ");"
    )
    for allocation in preview["cost_allocations"]:
        lines.extend(
            [
                "INSERT INTO gw_payroll_cost_allocations (payroll_period_id, workplace_id, allocation_key, amount, source_document_id, raw_payload)",
                (
                    "SELECT period.id, workplace.id, "
                    f"{sql_literal(allocation.get('allocation_key'))}, {allocation.get('amount', 0)}, "
                    f"(select id from gw_labor_source_documents where relative_path = {sql_literal(allocation.get('source_relative_path'))} order by created_at desc limit 1), "
                    f"{sql_json(allocation.get('raw_payload'))} "
                    "FROM gw_payroll_periods period "
                    "LEFT JOIN gw_workplaces workplace ON workplace.name = " + sql_literal(allocation.get("workplace_name")) + " "
                    f"WHERE {payroll_period_where(preview)}"
                ),
                "ON CONFLICT (payroll_period_id, workplace_id, allocation_key) DO UPDATE SET amount = excluded.amount, source_document_id = excluded.source_document_id, raw_payload = excluded.raw_payload;",
            ]
        )

    lines.extend(["", "NOTIFY pgrst, 'reload schema';", "COMMIT;", ""])
    return "\n".join(lines)


def build_preview(args: argparse.Namespace) -> dict[str, Any]:
    source_root = Path(args.source_root)
    payroll_kind = getattr(args, "payroll_kind", "monthly") or "monthly"
    documents = collect_documents(source_root, args.attendance_month, args.payroll_month, payroll_kind)
    extracted_pay_date = extract_pay_date_from_documents(documents, args.payroll_month)
    pay_date = extracted_pay_date or args.pay_date
    employees = extract_employee_master(documents)
    commute = extract_commute_data(source_root, documents)
    commute_monthly_results = aggregate_commute_monthly_results(commute["monthly_results"])
    checklist_rows = extract_payroll_checklists(source_root, documents)
    allocations = extract_cost_allocations(documents)

    known_names = {normalize_for_user_match(employee["display_name"]) for employee in employees}
    for employee in commute["extra_employees"]:
        key = normalize_for_user_match(employee["display_name"])
        if key and key not in known_names:
            employees.append(
                {
                    "employee_code": None,
                    "display_name": employee["display_name"],
                    "real_name": employee["real_name"],
                    "payroll_status": "active",
                    "raw_payload": employee["raw_payload"],
                }
            )
            known_names.add(key)

    payroll_results = extract_payroll_results(documents, employees, payroll_kind)
    for result in payroll_results:
        key = normalize_for_user_match(result.get("name"))
        if key and key not in known_names:
            employees.append(
                {
                    "employee_code": normalize_employee_code(result.get("employee_code")),
                    "display_name": result["name"],
                    "real_name": result["name"],
                    "payroll_status": "active",
                    "raw_payload": {
                        "source": "payroll_result",
                        "source_relative_path": result.get("source_relative_path"),
                        "page_number": result.get("page_number"),
                    },
                }
            )
            known_names.add(key)

    return {
        "source_root": str(source_root),
        "payroll_kind": payroll_kind,
        "attendance_month": args.attendance_month,
        "payroll_month": args.payroll_month,
        "period_start": args.attendance_month,
        "period_end": args.period_end or month_end(args.attendance_month),
        "pay_date": pay_date,
        "pay_date_source": "document" if extracted_pay_date else "default",
        "documents": documents,
        "employees": employees,
        "commute_routes": commute["routes"],
        "commute_monthly_results": commute_monthly_results,
        "payroll_checklists": checklist_rows,
        "cost_allocations": allocations,
        "payroll_results": payroll_results,
        "summary": {
            "payroll_kind": payroll_kind,
            "document_count": len(documents),
            "source_row_count": sum(len(document["rows"]) for document in documents),
            "employee_count": len(employees),
            "payroll_result_count": len(payroll_results),
            "payroll_result_payment_total": sum(result.get("payment_total") or 0 for result in payroll_results),
            "payroll_result_net_payment": sum(result.get("net_payment") or 0 for result in payroll_results),
            "commute_route_count": len(commute["routes"]),
            "commute_monthly_result_count": len(commute_monthly_results),
            "payroll_checklist_count": len(checklist_rows),
            "cost_allocation_count": len(allocations),
            "image_only_documents": [
                document["relative_path"]
                for document in documents
                if document["extraction_status"] == "image_only"
            ],
        },
    }


def build_output(preview: dict[str, Any]) -> dict[str, Any]:
    output = {
        key: value
        for key, value in preview.items()
        if key != "documents"
    }
    output["documents"] = [
        {key: value for key, value in document.items() if key != "rows"} | {"row_count": len(document["rows"])}
        for document in preview["documents"]
    ]
    return output


def find_batch_months(args: argparse.Namespace) -> list[dict[str, str]]:
    batch_root = Path(args.batch_root)
    if not batch_root.exists():
        raise FileNotFoundError(f"Batch root not found: {batch_root}")

    candidates: list[dict[str, str]] = []
    for child in sorted(batch_root.iterdir()):
        if not child.is_dir():
            continue
        payroll_month = payroll_month_from_folder_name(child.name)
        if not payroll_month:
            continue
        if not any(path.suffix.lower() in {".pdf", ".xlsx", ".xls"} for path in child.rglob("*")):
            continue
        candidates.append(
            {
                "source_root": str(child),
                "payroll_month": payroll_month,
                "payroll_kind": payroll_kind_from_folder_name(child.name),
                "attendance_month": add_months(payroll_month, -1),
                "pay_date": pay_date_for_month(payroll_month),
            }
        )

    candidates.sort(key=lambda item: (item["payroll_month"], 0 if item["payroll_kind"] == "monthly" else 1, item["source_root"]))
    if args.latest_payroll_month:
        candidates = [item for item in candidates if item["payroll_month"] <= args.latest_payroll_month]
    if args.months and args.months > 0:
        candidates = candidates[-args.months :]
    return candidates


def run_batch(args: argparse.Namespace) -> int:
    months = find_batch_months(args)
    if not months:
        print("No monthly folders found. Use names like 2026.01 or R08.01 under --batch-root.", file=sys.stderr)
        return 1

    json_out_dir = Path(args.json_out_dir or ".codex-tmp/labor-import-batch/json")
    sql_out_dir = Path(args.sql_out_dir or ".codex-tmp/labor-import-batch/sql")
    json_out_dir.mkdir(parents=True, exist_ok=True)
    sql_out_dir.mkdir(parents=True, exist_ok=True)

    manifest: list[dict[str, Any]] = []
    for item in months:
        namespace = argparse.Namespace(
            source_root=item["source_root"],
            attendance_month=item["attendance_month"],
            payroll_month=item["payroll_month"],
            payroll_kind=item["payroll_kind"],
            period_end=None,
            pay_date=item["pay_date"],
        )
        preview = build_preview(namespace)
        output = build_output(preview)
        stamp = item["payroll_month"][:7] if item["payroll_kind"] == "monthly" else f"{item['payroll_month'][:7]}_{item['payroll_kind']}"
        json_path = json_out_dir / f"{stamp}.json"
        sql_path = sql_out_dir / f"{stamp}.sql"
        json_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
        sql_path.write_text(generate_data_sql(preview), encoding="utf-8")
        manifest.append(
            {
                **item,
                "pay_date": preview["pay_date"],
                "pay_date_source": preview.get("pay_date_source", "default"),
                "json_path": str(json_path),
                "sql_path": str(sql_path),
                "summary": output["summary"],
            }
        )

    manifest_path = Path(args.batch_manifest or ".codex-tmp/labor-import-batch/manifest.json")
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps({"months": len(manifest), "manifest": str(manifest_path), "items": manifest}, ensure_ascii=False, indent=2))
    if args.apply:
        print("--apply is not used for batch mode. Use generated SQL files with Supabase CLI.", file=sys.stderr)
        return 1
    print("Batch dry-run only. Run each sql_path with: npx supabase db query --linked --file <sql_path> --output table")
    return 0


def sync_to_supabase(preview: dict[str, Any], client: SupabaseRest) -> dict[str, Any]:
    users = client.select("gw_users", "select=id,display_name,real_name,department,status")
    workplaces = client.select("gw_workplaces", "select=id,code,name")
    existing_employees = client.select(
        "gw_payroll_employees",
        "select=id,user_id,employee_code,display_name,real_name",
    )
    user_by_name: dict[str, dict[str, Any]] = {}
    for user in users:
        for name in (user.get("real_name"), user.get("display_name")):
            key = normalize_for_user_match(name)
            if key and key not in user_by_name:
                user_by_name[key] = user

    workplace_by_code = {workplace["code"]: workplace for workplace in workplaces}
    workplace_by_name = {workplace["name"]: workplace for workplace in workplaces}
    employee_by_code = {row.get("employee_code"): row for row in existing_employees if row.get("employee_code")}
    employee_by_user = {row.get("user_id"): row for row in existing_employees if row.get("user_id")}
    employee_by_name = {normalize_for_user_match(row.get("real_name") or row.get("display_name")): row for row in existing_employees}

    batch = client.insert(
        "gw_labor_import_batches",
        {
            "source_root": preview["source_root"],
            "payroll_kind": preview.get("payroll_kind", "monthly"),
            "target_attendance_month": preview["attendance_month"],
            "target_payroll_month": preview["payroll_month"],
            "period_start": preview["period_start"],
            "period_end": preview["period_end"],
            "pay_date": preview["pay_date"],
            "summary": preview["summary"],
        },
    )[0]

    attendance_period = client.upsert(
        "gw_attendance_periods",
        [
            {
                "attendance_month": preview["attendance_month"],
                "period_start": preview["period_start"],
                "period_end": preview["period_end"],
                "cutoff_day": int(preview["period_end"].split("-")[-1]),
                "rounding_unit_minutes": ATTENDANCE_ROUNDING_UNIT_MINUTES,
                "break_rule_set": "default",
            }
        ],
        "attendance_month",
    )[0]

    payroll_period = client.upsert(
        "gw_payroll_periods",
        [
            {
                "payroll_month": preview["payroll_month"],
                "payroll_kind": preview.get("payroll_kind", "monthly"),
                "attendance_period_id": attendance_period["id"],
                "attendance_month": preview["attendance_month"],
                "period_start": preview["period_start"],
                "period_end": preview["period_end"],
                "pay_date": preview["pay_date"],
            }
        ],
        "payroll_month,payroll_kind",
    )[0]

    document_id_by_path: dict[str, str] = {}
    for document in preview["documents"]:
        rows = document.pop("rows")
        payload = {**document, "import_batch_id": batch["id"]}
        synced = client.upsert("gw_labor_source_documents", [payload], "sha256")[0]
        document_id_by_path[document["relative_path"]] = synced["id"]
        client.delete("gw_labor_source_rows", f"source_document_id=eq.{synced['id']}")
        source_rows = []
        for row in rows:
            source_rows.append({**row, "source_document_id": synced["id"]})
        for chunk in chunked(source_rows, 100):
            client.insert("gw_labor_source_rows", chunk)

    employee_id_by_name: dict[str, str] = {}
    for employee in preview["employees"]:
        name_key = normalize_for_user_match(employee.get("real_name") or employee.get("display_name"))
        matched_user = user_by_name.get(name_key)
        target = None
        if employee.get("employee_code"):
            target = employee_by_code.get(employee["employee_code"])
        if not target and matched_user:
            target = employee_by_user.get(matched_user["id"])
        if not target:
            target = employee_by_name.get(name_key)

        workplace = None
        if matched_user and matched_user.get("department"):
            if matched_user["department"] == "フロア":
                workplace = workplace_by_code.get("aizu_brandhall")
            elif matched_user["department"] == "道の駅":
                workplace = workplace_by_code.get("michinoeki")
            else:
                workplace = workplace_by_code.get("hq")

        payload = {
            "display_name": employee["display_name"],
            "real_name": employee.get("real_name") or employee["display_name"],
            "kana": employee.get("kana"),
            "birth_date": employee.get("birth_date"),
            "hire_date": employee.get("hire_date"),
            "gender": employee.get("gender") or "unknown",
            "department": matched_user.get("department") if matched_user else None,
            "default_workplace_id": workplace.get("id") if workplace else None,
            "payroll_status": employee.get("payroll_status", "active"),
            "raw_payload": employee.get("raw_payload", {}),
        }
        if employee.get("employee_code"):
            payload["employee_code"] = employee["employee_code"]
        if matched_user:
            payload["user_id"] = matched_user["id"]

        if target:
            updated = client.patch("gw_payroll_employees", f"id=eq.{target['id']}", payload)[0]
        else:
            updated = client.insert("gw_payroll_employees", payload)[0]

        employee_id_by_name[name_key] = updated["id"]
        if updated.get("employee_code"):
            employee_by_code[updated["employee_code"]] = updated

    employee_rows = client.select("gw_payroll_employees", "select=id,user_id,employee_code,display_name,real_name")
    employee_id_by_name.update(
        {
            normalize_for_user_match(row.get("real_name") or row.get("display_name")): row["id"]
            for row in employee_rows
        }
    )

    route_rows = []
    for route in preview["commute_routes"]:
        employee_id = employee_id_by_name.get(normalize_for_user_match(route.pop("name")))
        workplace = workplace_by_code.get(route.pop("workplace_code"))
        source_path = route.pop("source_relative_path", None)
        if not employee_id:
            continue
        route_rows.append(
            {
                **route,
                "employee_id": employee_id,
                "workplace_id": workplace.get("id") if workplace else None,
                "source_document_id": document_id_by_path.get(source_path) if source_path else None,
            }
        )
    for chunk in chunked(route_rows):
        client.upsert("gw_commute_routes", chunk, "employee_id,workplace_id,route_type,effective_from")

    checklist_rows = []
    for checklist in preview["payroll_checklists"]:
        workplace = workplace_by_name.get(checklist.pop("workplace_name"))
        source_path = checklist.pop("source_relative_path", None)
        checklist_rows.append(
            {
                **checklist,
                "payroll_period_id": payroll_period["id"],
                "workplace_id": workplace.get("id") if workplace else None,
                "source_document_id": document_id_by_path.get(source_path) if source_path else None,
            }
        )
    for chunk in chunked(checklist_rows):
        client.insert("gw_payroll_checklists", chunk)

    allocation_rows = []
    for allocation in preview["cost_allocations"]:
        workplace = workplace_by_name.get(allocation.pop("workplace_name"))
        source_path = allocation.pop("source_relative_path", None)
        allocation_rows.append(
            {
                **allocation,
                "payroll_period_id": payroll_period["id"],
                "workplace_id": workplace.get("id") if workplace else None,
                "source_document_id": document_id_by_path.get(source_path) if source_path else None,
            }
        )
    for chunk in chunked(allocation_rows):
        client.upsert("gw_payroll_cost_allocations", chunk, "payroll_period_id,workplace_id,allocation_key")

    return {
        "batch_id": batch["id"],
        "attendance_period_id": attendance_period["id"],
        "payroll_period_id": payroll_period["id"],
        "document_count": len(document_id_by_path),
        "employee_count": len(employee_id_by_name),
        "commute_route_count": len(route_rows),
        "payroll_checklist_count": len(checklist_rows),
        "cost_allocation_count": len(allocation_rows),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", default=str(DEFAULT_SOURCE_ROOT))
    parser.add_argument("--batch-root", default="")
    parser.add_argument("--months", type=int, default=0)
    parser.add_argument("--latest-payroll-month", default="")
    parser.add_argument("--attendance-month", default=DEFAULT_ATTENDANCE_MONTH)
    parser.add_argument("--payroll-month", default=DEFAULT_PAYROLL_MONTH)
    parser.add_argument("--payroll-kind", default="monthly", choices=sorted(PAYROLL_KIND_LABELS))
    parser.add_argument("--period-end", default=None)
    parser.add_argument("--pay-date", default=DEFAULT_PAY_DATE)
    parser.add_argument("--env-file", default=".env.local")
    parser.add_argument("--json-out", default="")
    parser.add_argument("--sql-out", default="")
    parser.add_argument("--json-out-dir", default="")
    parser.add_argument("--sql-out-dir", default="")
    parser.add_argument("--batch-manifest", default="")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    if args.batch_root:
        return run_batch(args)

    preview = build_preview(args)
    output = build_output(preview)

    if args.json_out:
        json_out = Path(args.json_out)
        json_out.parent.mkdir(parents=True, exist_ok=True)
        json_out.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    if args.sql_out:
        sql_out = Path(args.sql_out)
        sql_out.parent.mkdir(parents=True, exist_ok=True)
        sql_out.write_text(generate_data_sql(preview), encoding="utf-8")

    print(json.dumps(output["summary"], ensure_ascii=False, indent=2))

    if not args.apply:
        print("Dry-run only. Add --apply to insert/upsert into Supabase, or run --sql-out with supabase db query.")
        return 0

    env = {**read_env_file(Path(args.env_file)), **os.environ}
    url = env.get("NEXT_PUBLIC_SUPABASE_URL", "")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        raise RuntimeError("NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required for --apply")

    result = sync_to_supabase(preview, SupabaseRest(url, key))
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
